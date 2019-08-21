import openpyxl as xl
import re
from openpyxl.chart import BarChart, Reference


class dbm_instrument:
    def __init__(self, inst_tag_name, inst_description, inst_io_type, inst_location):
        self.inst_tag_name = inst_tag_name
        self.inst_description = inst_description
        self.inst_io_type = inst_io_type
        self.inst_location = inst_location
        self.all = {
            'inst_tag_name': self.inst_tag_name,
            'inst_description': self.inst_description,
            'inst_io_type': self.inst_io_type,
            'inst_location': self.inst_location,
        }


class dbm_junction_box:
    def __init__(self, *args, **kwargs):
        for arg in args:
            print(arg)
            if arg == "generic":
                self.div = [9]
                self.dov = [5]
                self.aiv = [2]
                self.aic = [2]
                self.aov = [3]
                self.aoc = [3]
                self.dual_aiv = [2]
                self.dual_aic = [0]
                self.cameras = [2]
                self.lube_pump = [1]
                self.water_ingress = [1]
        for k, v in kwargs.items():
            print(k, v)
            if k == 'name':
                self.name = v


def create_instruments(input_filename):
    allIO = []
    wb = xl.load_workbook(input_filename, data_only=True)
    sheet = wb['Instrument List']
    for row in range(19, sheet.max_row + 1):
        # Try to get the column range in the row to check which io type it is and record tag_io_type as such
        """ row_range = str("P{}:V{}".format(row, row))
        print(sheet.cell(row_range)) """

        if str(sheet.cell(row, 5).value).startswith("C1"):
            inst_tag_name = sheet.cell(row, 5).value
            inst_description = str.strip(sheet.cell(row, 6).value)
            inst_location = str.strip(sheet.cell(row, 7).value)
            inst_io_type = sheet.cell(row, 16).value
            # remove any whitespace
            cleaned_tag_name = re.sub('[\s+]', '', inst_tag_name)
            # print(cleaned_cell)
            ob = dbm_instrument(
                cleaned_tag_name, inst_description, inst_io_type, inst_location)
            ob_obj = {
                "inst_tag_name": cleaned_tag_name,
                "inst_description": inst_description,
                "inst_io_type": inst_io_type,
                "inst_location": inst_location
            }
            # print(ob.tag_description)
            allIO.append(ob)
    print(allIO[1].all)
    sort_all_io = sorted(
        allIO, key=lambda x: x.inst_tag_name, reverse=False)
    for i in sort_all_io:
        print(i.inst_tag_name)
    # print(allIO[1].all)


create_instruments('inst_list.xlsx')

t_jb = dbm_junction_box('generic', name="Test_JB")
print(t_jb.aic)
print(t_jb.name)
