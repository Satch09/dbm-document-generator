import openpyxl as xl
from openpyxl.chart import BarChart, Reference


def process_workbook(input_filename, output_filename):

    wb = xl.load_workbook(input_filename)
    sheet = wb['Sheet1']
    cell = sheet.cell(2, 1)

    for row in range(2, sheet.max_row + 1):
        print(sheet.cell(row, 5).value)
        cell = sheet.cell(row, 5)
        correct_price = cell.value * 1.14
        correct_price_cell = sheet.cell(row, 6)
        correct_price_cell.value = correct_price
    values = Reference(sheet,
                       min_row=2,
                       max_row=sheet.max_row,
                       min_col=6,
                       max_col=6
                       )
    chart = BarChart()
    chart.add_data(values)
    sheet.add_chart(chart, 'H1')
    wb.save(output_filename)


with open('')
