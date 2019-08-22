import openpyxl as xl
import tkinter as Tk
from instrument import Instrument
from io_card import IO_Card
from io_rack import IO_Rack

wb = xl.load_workbook('MRC_Inst_List.xlsx', data_only=True)
sheet = wb['data']


def get_instruments(data):
    item = {}
    instruments = []
    for row in range(2, data.max_row):
        for col in range(1, data.max_column):
            key = data.cell(1, col).value
            value = data.cell(row, col).value
            result = {key: value}
            item.update(result)
        instruments.append(Instrument(item))
    return instruments


my_insts = get_instruments(sheet)


def create_racks(instruments, **params):
    """
    params = 'DIV' = '1756-IB32', 'DOV' = '1756-OB32',
    """
    for p in params.keys():
        q = []
        for instrument in instruments:
            if instrument.inst_io_type == p:
                q.append(instrument)
        print('-' * 50)
        crd = IO_Card(params.values(), p, my_insts[:32])
        print(p)
        for i in q:
            print("{}: {}".format(i, i.tag_description))
    print(crd)

# firstRange = my_insts[:32]
# secondRange = my_insts[33:65]
# thirdRange = my_insts[66:98]
# forthRange = my_insts[99:]

# c1 = IO_Card('1756-IB32', 'DIV')
# c1.populate_card(
#     (firstRange))
# c2 = IO_Card('1756-IB32', 'DIV')
# c2.populate_card((secondRange))
# c3 = IO_Card('1756-IB32', 'DIV')
# c3.populate_card((thirdRange))
# c4 = IO_Card('1756-IB32', 'DIV')
# c4.populate_card((forthRange))
# rack = IO_Rack(4, [c1, c2, c3, c4])
# rack = IO_Rack(4, test)
# print(str(rack.node_number))
# print(str(rack.rack_number))
# print(str(rack.slot_number))
# for r in rack.rack:
#     print('-' * 100)
#     print(r.part_number)
#     for s in r.io_assignments:
#         print("{}: {}".format(s, s.tag_description))


create_racks(my_insts, AIC='1756-IF16')
